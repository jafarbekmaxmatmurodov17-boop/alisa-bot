"""
ALISA TELEGRAM BOT v6.0
========================
Ko'p hujjat qo'llab-quvvatlash:
- Bir vaqtda Invoice + Packing list + CMR + Sertifikat yuborilsa
- Hammasi birlashtirib tahlil qilinadi
- 40+ grafa to'ldiriladi
"""

import os, re, logging, tempfile, requests, base64, json
from pathlib import Path
from datetime import datetime
from io import BytesIO
from collections import defaultdict

from telegram import Update, Message
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters

import fitz
from PIL import Image

try:
    from docx import Document as DocxDoc
    DOCX_OK = True
except: DOCX_OK = False

try:
    import openpyxl
    EXCEL_OK = True
except: EXCEL_OK = False

BOT_TOKEN  = "8955671081:AAFmM4eiWeCmZpuYsSEg_MIPF6B1TOeJHbU"
GEMINI_KEY = "AIzaSyCR64mprAkkkfzmYw2lELjVTHbz8QybF8w"
MODEL      = "gemini-2.5-flash"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1/models/{MODEL}:generateContent?key={GEMINI_KEY}"

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

# Guruh bo'yicha hujjatlarni to'plash (albom/media group uchun)
pending_groups: dict = {}  # media_group_id → list of (path, filename)
pending_tasks: dict = {}   # chat_id → list of (path, filename) [alohida fayllar uchun]

BYD_GRAFALAR = {
    "1":"Deklaratsiya turi","2":"Jo'natuvchi/Eksportyor","3":"Shakl",
    "4":"Yukxatlar","5":"Tovar turlari soni","6":"O'rin soni",
    "7":"Hujjat raqami","8":"Qabul qiluvchi/Importyor",
    "9":"Moliyaviy javobgar","10":"Jo'natuvchi mamlakat",
    "11":"Savdo mamlakati","12":"Qiymat","13":"GTD raqami",
    "14":"Deklarant","15":"Jo'natish mamlakati",
    "16":"Kelib chiqish mamlakati","17":"Belgilangan mamlakat",
    "18":"Transport","19":"Konteyner","20":"Incoterms",
    "21":"Chegarada transport","22":"Valyuta va summa",
    "23":"Valyuta kursi","24":"Bitim turi","25":"Transport (chegara)",
    "26":"Ichki transport","27":"Yuklash joyi","28":"Bank",
    "29":"Kirish bojxonasi","30":"Tovar joylashuvi",
    "31":"Tovar tavsifi","32":"Tartib raqami","33":"HS kod",
    "34":"Kelib chiqish kodi","35":"Yalpi og'irlik kg",
    "36":"Imtiyozlar","37":"Tartib-taomil","38":"Sof og'irlik kg",
    "39":"Kvota","40":"Oldingi hujjat","41":"O'lchov/miqdor",
    "42":"Tovar qiymati","43":"QK usuli","44":"Hujjatlar",
    "45":"Tuzatish","46":"Statistik qiymat","47":"To'lovlar",
    "48":"Kechiktirilgan to'lov","49":"Ombor","50":"Majburiyatli shaxs",
    "51":"O'tish bojxonalari","52":"Kafolat","53":"Boshqaruv bojxonasi",
    "54":"Joy, sana, imzo",
}

SYSTEM_PROMPT = """Siz O'zbekiston Respublikasi Davlat bojxona qo'mitasining 2773-sonli (06.04.2016) 
"Bojxona Yuk Deklaratsiyasini to'ldirish tartibi to'g'risidagi Yo'riqnoma" bo'yicha mutaxassisiz.

Sizga bir yoki bir nechta hujjat matni beriladi (Invoice, Packing list, CMR, Xitoy sertifikati va h.k.)
Siz ulardan BYD (Bojxona Yuk Deklaratsiyasi) grafalarini aniqlab, JSON formatida qaytarasiz.

MUHIM: FAQAT JSON qaytaring. Boshqa hech narsa yozmang.
Format: {"1": "qiymat", "2": "qiymat", ...}

=== 2773-SON YO'RIQNOMA BO'YICHA GRAFALAR ===

1=Deklaratsiya turi: "IM 40 00" (import), "EX 10 11" (eksport), "TR 80 00" (tranzit)
2=Jo'natuvchi/Eksportyor: Seller/Shipper/Exporter/卖家 — nomi va manzili
3=Shakl: "1/1" (1 tovar, 1 varaq)
4=Yukxatlar soni: ilova hujjatlar soni
5=Tovar turlari soni: Invoice da pozitsiyalar soni
6=O'rin soni: Packages/Cartons/PCS/件/Мест soni
7=Hujjat raqami: sertifikat raqami (合格证编号) yoki ichki raqam
8=Qabul qiluvchi/Importyor: Buyer/Consignee/ПОКУПАТЕЛЬ — nomi va manzili
10=Jo'natuvchi mamlakat: Country of Shipment (CN, DE, KR...)
11=Savdo mamlakati: tovar sotib olingan mamlakat
15=Jo'natish mamlakati: 2 harfli ISO (CN, DE...)
16=Kelib chiqish mamlakati: Country of Origin/原产地 (CN...)
17=Belgilangan mamlakat: import uchun "UZ"
18=Transport vositasi: CMR/TIR da transport raqami va mamlakati
19=Konteyner: "1" (bor) yoki "0" (yo'q)
20=Incoterms: CIP/FOB/CIF/EXW + shahar (kontraktdan)
21=Chegarada transport turi: "3"=avtomobil, "2"=temir yo'l, "4"=avia
22=Valyuta va summa: "USD 19800.00" (Invoice Total)
24=Bitim turi: "11" (oddiy sotib olish-sotish)
25=Transport turi chegarada: raqam kodi
27=Yuklash joyi: Port of Loading / yuklash punkti (CMR dan)
31=Tovar tavsifi: nomi, markasi, modeli, VIN (17 ta belgi), texnik xarakteristika
33=HS kod: 10 xonali TN VED kodi (avtomobil: 8703...)
34=Kelib chiqish kodi: 2 harfli (CN, DE, KR, JP...)
35=Yalpi og'irlik kg: Gross Weight (Packing list / 总质量)
38=Sof og'irlik kg: Net Weight (Packing list / 整备质量)
40=Oldingi hujjat: CMR raqami yoki TIR Carnet raqami
41=O'lchov/miqdor: "1 dona" yoki "X sht"
42=Tovar qiymati: Invoice da narx (valyutada)
43=QK usuli: "1" (bitim qiymati)
44=Hujjatlar: Invoice №..., Packing list №..., Contract №..., CMR №...
54=Sana: deklaratsiya yoki Invoice sanasi

=== HUJJAT TURLARIDAN MA'LUMOT ===
INVOICE: 2→Seller, 8→Buyer, 22→Total, 20→Incoterms, 31→Description, 41→Qty, 42→Price, 44→Invoice No
PACKING LIST: 6→Packages, 35→Gross Weight, 38→Net Weight
CMR: 18→Transport plate, 27→Loading point, 40→CMR No
合格证 (Xitoy sertifikati): 7→合格证编号, 31→车辆品牌+车型型号+VIN, 35→总质量, 38→整备质量, 54→发证日期

Hujjatda aniq mavjud ma'lumotlarni kiriting. Bo'sh qolsa — o'sha grafani kiritma.
FAQAT JSON qaytaring."""


def parse_json(text: str) -> dict:
    text = re.sub(r'```[a-zA-Z]*', '', text).strip().strip('`').strip()
    try:
        r = json.loads(text)
        if isinstance(r, dict): return r
    except: pass
    m = re.search(r'\{[\s\S]*\}', text)
    if m:
        try:
            r = json.loads(m.group(0))
            if isinstance(r, dict): return r
        except: pass
    pairs = re.findall(r'"(\d+)"\s*:\s*"([^"]*)"', text)
    return dict(pairs) if pairs else {}


def gemini_text(prompt: str) -> dict:
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.05, "maxOutputTokens": 4096}
    }
    try:
        r = requests.post(GEMINI_URL, json=payload, timeout=120)
        if r.status_code != 200:
            log.error(f"Gemini {r.status_code}: {r.text[:300]}")
            return {}
        data = r.json()
        answer = data["candidates"][0]["content"]["parts"][0]["text"]
        log.info(f"Gemini javob ({len(answer)} belgi): {answer[:150]}")
        return parse_json(answer)
    except Exception as e:
        log.error(f"Gemini xato: {e}")
        return {}


def gemini_vision(b64: str, mime: str, extra_text: str = "") -> dict:
    prompt = SYSTEM_PROMPT + f"\n\n{extra_text}\nRasmdagi hujjatdan BYD grafalarini aniqlab JSON qaytaring:"
    payload = {
        "contents": [{"parts": [
            {"text": prompt},
            {"inlineData": {"mimeType": mime, "data": b64}}
        ]}],
        "generationConfig": {"temperature": 0.05, "maxOutputTokens": 4096}
    }
    try:
        r = requests.post(GEMINI_URL, json=payload, timeout=120)
        if r.status_code != 200:
            log.error(f"Vision {r.status_code}: {r.text[:300]}")
            return {}
        data = r.json()
        answer = data["candidates"][0]["content"]["parts"][0]["text"]
        return parse_json(answer)
    except Exception as e:
        log.error(f"Vision xato: {e}")
        return {}


def merge_fields(base: dict, new: dict) -> dict:
    """Ikki natijani birlashtiradi — bo'sh bo'lsa yangi qiymat oladi."""
    result = dict(base)
    for k, v in new.items():
        if v and str(v).strip():
            if k not in result or not result[k]:
                result[k] = v
            elif k == "44":  # Hujjatlar ro'yxatini birlashtirish
                existing = result[k]
                if str(v) not in existing:
                    result[k] = existing + ", " + str(v)
            elif k == "31":  # Tovar tavsifini birlashtirish
                existing = result[k]
                if str(v) not in existing:
                    result[k] = existing + " | " + str(v)
    return result


# ── Fayl o'qish ──

def pdf_pages(path: str):
    imgs = []
    try:
        doc = fitz.open(path)
        for page in doc:
            mat = fitz.Matrix(2.5, 2.5)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.open(BytesIO(pix.tobytes("png")))
            imgs.append(img)
        doc.close()
    except Exception as e:
        log.error(f"PDF pages: {e}")
    return imgs


def pdf_text(path: str) -> str:
    text = ""
    try:
        doc = fitz.open(path)
        for page in doc:
            text += page.get_text("text") + "\n"
        doc.close()
    except: pass
    return text.strip()


def docx_text(path: str) -> str:
    if not DOCX_OK: return ""
    try:
        doc = DocxDoc(path)
        lines = [p.text for p in doc.paragraphs if p.text.strip()]
        for table in doc.tables:
            for row in table.rows:
                vals = [c.text.strip() for c in row.cells if c.text.strip()]
                if vals: lines.append(" | ".join(vals))
        return "\n".join(lines)
    except: return ""


def excel_text(path: str) -> str:
    if not EXCEL_OK: return ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                vals = [str(c) for c in row if c and str(c).strip()]
                if vals: lines.append(" | ".join(vals))
        return "\n".join(lines)
    except: return ""


def img_b64(img: Image.Image) -> str:
    buf = BytesIO()
    if img.mode != "RGB": img = img.convert("RGB")
    img.save(buf, format="JPEG", quality=90)
    return base64.b64encode(buf.getvalue()).decode()


async def process_one_file(path: str, filename: str) -> dict:
    """Bitta faylni tahlil qiladi."""
    ext = Path(filename).suffix.lower()

    if ext in (".docx", ".doc"):
        text = docx_text(path)
        return gemini_text(SYSTEM_PROMPT + f"\n\nHUJJAT ({filename}):\n{text[:12000]}") if text else {}

    elif ext in (".xlsx", ".xls"):
        text = excel_text(path)
        return gemini_text(SYSTEM_PROMPT + f"\n\nHUJJAT ({filename}):\n{text[:12000]}") if text else {}

    elif ext in (".txt", ".csv"):
        text = open(path, encoding="utf-8", errors="ignore").read()
        return gemini_text(SYSTEM_PROMPT + f"\n\nHUJJAT ({filename}):\n{text[:12000]}") if text else {}

    elif ext == ".pdf":
        text = pdf_text(path)
        if len(text.strip()) > 100:
            return gemini_text(SYSTEM_PROMPT + f"\n\nHUJJAT ({filename}):\n{text[:12000]}")
        else:
            fields = {}
            for i, img in enumerate(pdf_pages(path)[:4]):
                result = gemini_vision(img_b64(img), "image/jpeg", f"PDF sahifa {i+1}, fayl: {filename}")
                fields = merge_fields(fields, result)
                if len(fields) >= 15: break
            return fields

    elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp"):
        mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
        b64 = base64.b64encode(open(path, "rb").read()).decode()
        return gemini_vision(b64, mime, f"Fayl: {filename}")

    return {}


async def process_multiple_files(files: list, update: Update) -> dict:
    """
    Ko'p hujjatni birlashtirib tahlil qiladi.
    files = [(path, filename), ...]
    """
    all_texts = []
    image_parts = []
    fields = {}

    await update.message.reply_text(
        f"📦 *{len(files)} ta hujjat* qabul qilindi.\n"
        f"🤖 Barchasi birgalikda tahlil qilinmoqda...",
        parse_mode="Markdown"
    )

    for path, filename in files:
        ext = Path(filename).suffix.lower()
        log.info(f"Qayta ishlash: {filename}")

        if ext in (".docx", ".doc"):
            text = docx_text(path)
            if text: all_texts.append(f"=== {filename} ===\n{text[:6000]}")

        elif ext in (".xlsx", ".xls"):
            text = excel_text(path)
            if text: all_texts.append(f"=== {filename} ===\n{text[:6000]}")

        elif ext in (".txt", ".csv"):
            text = open(path, encoding="utf-8", errors="ignore").read()
            if text: all_texts.append(f"=== {filename} ===\n{text[:6000]}")

        elif ext == ".pdf":
            text = pdf_text(path)
            if len(text.strip()) > 100:
                all_texts.append(f"=== {filename} ===\n{text[:6000]}")
            else:
                pages = pdf_pages(path)
                for i, img in enumerate(pages[:2]):
                    image_parts.append((img_b64(img), "image/jpeg", filename))

        elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".webp"):
            mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
            b64 = base64.b64encode(open(path, "rb").read()).decode()
            image_parts.append((b64, mime, filename))

    # Matnli hujjatlarni birgalikda yuborish
    if all_texts:
        combined = "\n\n".join(all_texts)
        prompt = (SYSTEM_PROMPT +
                  f"\n\nQuyida {len(all_texts)} ta hujjat mavjud. "
                  f"Hammasidan BYD grafalarini aniqlab JSON qaytaring:\n\n{combined[:18000]}")
        result = gemini_text(prompt)
        fields = merge_fields(fields, result)

    # Rasmli hujjatlarni alohida yuborish
    for b64, mime, fname in image_parts:
        result = gemini_vision(b64, mime, f"Fayl: {fname}")
        fields = merge_fields(fields, result)

    return fields


def format_byd(fields: dict, filenames: list) -> str:
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    fnames = ", ".join(f"`{f}`" for f in filenames[:3])
    if len(filenames) > 3: fnames += f" +{len(filenames)-3}"

    lines = [
        "📋 *ALISA v6.0 — BYD TAHLILI*",
        f"⚖️ 2773-son Yo'riqnoma asosida",
        f"🤖 Gemini 2.5 Flash  •  {len(filenames)} ta hujjat",
        f"📎 {fnames}  🕐 {now}",
        "─" * 35,
    ]

    valid = {k: str(v).strip() for k, v in fields.items()
             if v and str(v).strip() and k.isdigit()}

    if not valid:
        lines += [
            "⚠️ Hujjatlardan ma'lumot ajratib bo'lmadi.",
            "💡 PDF, Word, Excel yoki aniq JPG yuboring.",
        ]
        return "\n".join(lines)

    found = sorted(valid.keys(), key=lambda x: int(x))
    lines.append(f"✅ *{len(found)} ta grafa to'ldirildi:*\n")

    for num in found:
        name = BYD_GRAFALAR.get(num, f"Grafa {num}")
        val = valid[num]
        if len(val) > 150: val = val[:150] + "..."
        lines.append(f"*{num}* — _{name}_")
        lines.append(f"   `{val}`\n")

    missing = [n for n in BYD_GRAFALAR if n not in valid]
    if missing:
        lines.append("─" * 35)
        lines.append(f"⚠️ *Qo'lda to'ldirish kerak ({len(missing)} ta):*")
        for grp in [missing[i:i+8] for i in range(0, len(missing), 8)]:
            lines.append("  " + " · ".join(grp))

    return "\n".join(lines)


def eombor(query: str) -> str:
    try:
        r = requests.get(
            f"https://e-ombor.uz/search?q={requests.utils.quote(query)}",
            headers={"User-Agent": "Mozilla/5.0"}, timeout=8
        )
        codes = list(dict.fromkeys(re.findall(r'\b\d{8,10}\b', r.text)))[:5]
        return f"HS kodlar: {', '.join(codes)}" if codes else "Topilmadi"
    except Exception as e:
        return f"Xato: {e}"


async def download_file(ctx, file_id: str, suffix: str) -> str:
    tg_file = await ctx.bot.get_file(file_id)
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp_path = tmp.name
    await tg_file.download_to_drive(tmp_path)
    return tmp_path


# ── Handlers ──

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Salom! Men ALISA v6.0*\n"
        "⚖️ *2773-son Yo'riqnoma* asosida\n"
        "🤖 *Gemini 2.5 Flash AI*\n\n"
        "📎 *Ko'p hujjat yuboring:*\n"
        "   Invoice + Packing list + CMR\n"
        "   → Hammasi birlashtirib tahlil!\n\n"
        "✅ *Formatlar:* PDF · Word · Excel · JPG\n\n"
        "💡 *Maslahat:*\n"
        "Bir nechta hujjatni bir vaqtda yuboring\n"
        "(Paperclip → Bir nechta fayl tanlang)\n\n"
        "/tahlil — Yuborilgan fayllarni tahlil qilish\n"
        "/tozala — Fayllarni tozalash\n"
        "/eombor [tovar] — HS kod qidirish",
        parse_mode="Markdown"
    )

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📚 *ALISA v6.0 YORDAM*\n\n"
        "*Ko'p hujjat rejimi:*\n"
        "1. Invoice yuborilsa → 2,8,22,31,44-grafalar\n"
        "2. Packing list qo'shilsa → 6,35,38-grafalar\n"
        "3. CMR qo'shilsa → 18,27,40-grafalar\n"
        "4. 合格证 qo'shilsa → 7,31,35,38-grafalar\n"
        "Jami: *40+ grafa* to'ldiriladi!\n\n"
        "*Ishlatish:*\n"
        "Bir vaqtda bir nechta fayl yuboring\n"
        "Bot hammasi birlashib tahlil qiladi\n\n"
        "*/eombor elektromobil* — HS kod",
        parse_mode="Markdown"
    )

async def cmd_eombor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not ctx.args:
        await update.message.reply_text("Misol: `/eombor elektromobil`", parse_mode="Markdown")
        return
    q = " ".join(ctx.args)
    await update.message.reply_text(f"🔍 `{q}`...", parse_mode="Markdown")
    await update.message.reply_text(f"📦 {eombor(q)}")

async def cmd_tahlil(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """To'plangan fayllarni tahlil qilish."""
    chat_id = update.effective_chat.id
    files = pending_tasks.get(chat_id, [])
    if not files:
        await update.message.reply_text(
            "📭 Hujjat yo'q.\n"
            "Avval hujjatlarni yuboring, keyin /tahlil bosing."
        )
        return
    fields = await process_multiple_files(files, update)
    for path, _ in files:
        try: os.unlink(path)
        except: pass
    pending_tasks.pop(chat_id, None)
    reply = format_byd(fields, [f for _, f in files])
    if len(reply) > 4000: reply = reply[:4000] + "\n...(qisqartirildi)"
    await update.message.reply_text(reply, parse_mode="Markdown")

async def cmd_tozala(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    files = pending_tasks.pop(chat_id, [])
    for path, _ in files:
        try: os.unlink(path)
        except: pass
    await update.message.reply_text(f"🗑 {len(files)} ta fayl tozalandi.")

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc: return
    filename = doc.file_name or "hujjat"
    ext = Path(filename).suffix.lower()
    chat_id = update.effective_chat.id
    media_group_id = update.message.media_group_id

    try:
        tmp_path = await download_file(ctx, doc.file_id, ext or ".tmp")
    except Exception as e:
        await update.message.reply_text(f"❌ Yuklab bo'lmadi: {e}")
        return

    if media_group_id:
        # Albom (bir vaqtda ko'p fayl)
        if media_group_id not in pending_groups:
            pending_groups[media_group_id] = []
            # 3 soniyadan keyin tahlil qilish
            ctx.job_queue.run_once(
                process_group,
                3,
                data={"group_id": media_group_id, "chat_id": chat_id, "message": update.message},
                name=f"group_{media_group_id}"
            )
        pending_groups[media_group_id].append((tmp_path, filename))
        await update.message.reply_text(
            f"📎 `{filename}` qabul qilindi ({len(pending_groups[media_group_id])} ta)...",
            parse_mode="Markdown"
        )
    else:
        # Yagona fayl — darhol tahlil
        await update.message.reply_text(
            f"📎 *{filename}* tahlil qilinmoqda...\n"
            f"💡 Ko'p hujjat uchun /tahlil buyrug'ini ishlating",
            parse_mode="Markdown"
        )
        fields = await process_one_file(tmp_path, filename)
        try: os.unlink(tmp_path)
        except: pass
        reply = format_byd(fields, [filename])
        if len(reply) > 4000: reply = reply[:4000] + "\n...(qisqartirildi)"
        await update.message.reply_text(reply, parse_mode="Markdown")
        if "31" in fields and fields["31"]:
            r = eombor(str(fields["31"])[:50])
            if "HS kodlar" in r:
                await update.message.reply_text(f"🌐 e-ombor.uz: {r}")


async def process_group(context):
    """Albom fayllarini tahlil qilish."""
    data = context.job.data
    group_id = data["group_id"]
    chat_id = data["chat_id"]
    message = data["message"]

    files = pending_groups.pop(group_id, [])
    if not files:
        return

    fields = await process_multiple_files(files, message)
    for path, _ in files:
        try: os.unlink(path)
        except: pass

    reply = format_byd(fields, [f for _, f in files])
    if len(reply) > 4000: reply = reply[:4000] + "\n...(qisqartirildi)"
    await context.bot.send_message(chat_id, reply, parse_mode="Markdown")

    if "31" in fields and fields["31"]:
        r = eombor(str(fields["31"])[:50])
        if "HS kodlar" in r:
            await context.bot.send_message(chat_id, f"🌐 e-ombor.uz: {r}")


async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    photo = update.message.photo[-1]
    media_group_id = update.message.media_group_id
    chat_id = update.effective_chat.id

    try:
        tmp_path = await download_file(ctx, photo.file_id, ".jpg")
    except Exception as e:
        await update.message.reply_text(f"❌ {e}")
        return

    if media_group_id:
        if media_group_id not in pending_groups:
            pending_groups[media_group_id] = []
            ctx.job_queue.run_once(
                process_group, 3,
                data={"group_id": media_group_id, "chat_id": chat_id, "message": update.message},
                name=f"group_{media_group_id}"
            )
        filename = f"rasm_{len(pending_groups[media_group_id])+1}.jpg"
        pending_groups[media_group_id].append((tmp_path, filename))
        await update.message.reply_text(
            f"🖼 Rasm qabul qilindi ({len(pending_groups[media_group_id])} ta)...",
        )
    else:
        await update.message.reply_text("🖼 Gemini Vision tahlil qilmoqda...")
        fields = await process_one_file(tmp_path, "rasm.jpg")
        try: os.unlink(tmp_path)
        except: pass
        reply = format_byd(fields, ["rasm.jpg"])
        if len(reply) > 4000: reply = reply[:4000] + "\n...(qisqartirildi)"
        await update.message.reply_text(reply, parse_mode="Markdown")


async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    if any(w in text.lower() for w in ["salom", "привет", "hello"]):
        await cmd_start(update, ctx)
    else:
        await update.message.reply_text(
            "📎 Hujjat yuboring (PDF/JPG/Word/Excel)\n"
            "Bir vaqtda bir nechta → /tahlil\n"
            "`/eombor tovar` — HS kod",
            parse_mode="Markdown"
        )


def main():
    print("=" * 55)
    print(f"  ALISA BOT v6.0 — Ko'p hujjat + 2773-Yo'riqnoma")
    print("=" * 55)
    app = (Application.builder()
           .token(BOT_TOKEN)
           .build())

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("eombor", cmd_eombor))
    app.add_handler(CommandHandler("tahlil", cmd_tahlil))
    app.add_handler(CommandHandler("tozala", cmd_tozala))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    print("✅ Bot tayyor!")
    print("💡 Bir vaqtda ko'p fayl yuborish uchun Paperclip → ko'p fayl tanlang")
    print("=" * 55)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
